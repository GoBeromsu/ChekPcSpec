import os
import psutil
import platform
import openpyxl
import os.path
from pathlib import Path

process = psutil.swap_memory().total/1073741824 # 메모리 용량
cpu = platform.machine() # cpu 정보
cpu_num = psutil.cpu_count()
os = platform.system() # os 정보

# device 정보
index = 1
disk = psutil.disk_partitions()[0].device
disk_type = psutil.disk_partitions()[0].fstype
disk_auth = psutil.disk_partitions()[0].opts
disk_mount = psutil.disk_partitions()[0].mountpoint
disk_total = psutil.disk_usage(disk_mount).total/1073741824

title = ['Index','Os','CPU','cpu 갯수','메모리 용량','Disk','Disk type','Disk 용량','Disk 권한','Disk 마운트 위치']


if __name__ == "__main__":
    my_file = Path('./Pc_SPEC.xlsx')
    list = [index, os,cpu,cpu_num,process,disk,disk_type,disk_total,disk_auth,disk_mount]
    if my_file.is_file():
        wb = openpyxl.load_workbook(my_file)
        sheet = wb.active
        list[0] = sheet.max_column - 1
        sheet.append(list)
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(title)
        sheet.append(list)    
    
    wb.save('./Pc_SPEC.xlsx')
    wb.close()
